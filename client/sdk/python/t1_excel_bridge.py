"""
T1 Broker Excel Trading Bridge
================================
Connects the Excel trading template to the T1 Broker API.
Reads orders, executes trades, updates positions and results.

Usage:
    python t1_excel_bridge.py connect       # Test connection
    python t1_excel_bridge.py trade         # Execute batch orders from Order Entry sheet
    python t1_excel_bridge.py quick         # Execute single order from Quick Trade sheet
    python t1_excel_bridge.py refresh       # Refresh positions + order history
    python t1_excel_bridge.py instruments   # Load available instruments
    python t1_excel_bridge.py all           # Full sync (connect + instruments + refresh)

Requirements:
    pip install requests openpyxl
    t1_client.py must be in the same directory (or installed)
"""

import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from t1_client import T1Client, T1Error
except ImportError:
    # Try relative path
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    try:
        from t1_client import T1Client, T1Error
    except ImportError:
        print("ERROR: t1_client.py not found. Copy it to this directory.")
        print("  Location: client/sdk/python/t1_client.py")
        sys.exit(1)

# ── CONFIG ───────────────────────────────────────────────────────
TEMPLATE = "T1-Broker-Trading-Template.xlsx"

def find_template():
    """Find the Excel template file."""
    candidates = [TEMPLATE, f"./{TEMPLATE}", os.path.join(os.path.dirname(__file__), TEMPLATE)]
    for path in candidates:
        if os.path.exists(path):
            return path
    print(f"ERROR: '{TEMPLATE}' not found in current directory.")
    print(f"  Current directory: {os.getcwd()}")
    sys.exit(1)

def load_wb():
    path = find_template()
    return openpyxl.load_workbook(path), path

def save_wb(wb, path):
    wb.save(path)
    print(f"  Saved: {path}")

def get_client(wb):
    """Read API key and URL from Connection sheet, return T1Client."""
    ws = wb["Connection"]
    api_key = str(ws["C6"].value or "").strip()
    base_url = str(ws["C7"].value or "http://localhost:3000/api/v1").strip()
    timeout = int(ws["C8"].value or 30)

    if not api_key or not api_key.startswith("t1_live_"):
        print("ERROR: Invalid API key in Connection sheet (cell C6).")
        print("  Key must start with 't1_live_'")
        sys.exit(1)

    return T1Client(api_key, base_url=base_url, timeout=timeout)

def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ═══════════════════════════════════════════════════════════════════
# CONNECT — Test connection, fill Connection sheet status
# ═══════════════════════════════════════════════════════════════════
def cmd_connect():
    print("T1 Broker — Testing Connection...")
    wb, path = load_wb()
    client = get_client(wb)
    ws = wb["Connection"]

    try:
        is_up = client.ping()
        if not is_up:
            ws["C12"].value = "FAILED — Server not reachable"
            save_wb(wb, path)
            print("  FAILED: Server not reachable")
            return

        profile = client.profile()
        p = profile.get("data", profile)

        ws["C12"].value = "CONNECTED"
        ws["C12"].font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="3FB950")
        ws["C13"].value = p.get("email", p.get("user", {}).get("email", "—"))
        ws["C13"].font = openpyxl.styles.Font(name="Arial", size=10, color="E6EDF3")
        ws["C14"].value = p.get("role", p.get("user", {}).get("role", "—"))
        ws["C14"].font = openpyxl.styles.Font(name="Arial", size=10, color="E6EDF3")
        ws["C15"].value = p.get("kycStatus", p.get("kyc_status", "—"))
        ws["C15"].font = openpyxl.styles.Font(name="Arial", size=10, color="E6EDF3")

        trading = p.get("tradingEnabled", p.get("trading_enabled", None))
        ws["C16"].value = "YES" if trading else ("NO" if trading is not None else "—")
        ws["C16"].font = openpyxl.styles.Font(name="Arial", size=10, bold=True,
                                               color="3FB950" if trading else "F85149")

        # API key permissions
        keys = client.list_api_keys()
        if keys:
            perms = keys[0].get("permissions", [])
            ws["C17"].value = ", ".join(perms) if perms else "—"
        else:
            ws["C17"].value = "—"
        ws["C17"].font = openpyxl.styles.Font(name="Arial", size=10, color="E6EDF3")

        save_wb(wb, path)
        print(f"  CONNECTED as {ws['C13'].value} (role: {ws['C14'].value})")
        print(f"  KYC: {ws['C15'].value} | Trading: {ws['C16'].value}")

    except T1Error as e:
        ws["C12"].value = f"ERROR: {e}"
        ws["C12"].font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="F85149")
        save_wb(wb, path)
        print(f"  ERROR: {e}")

# ═══════════════════════════════════════════════════════════════════
# TRADE — Execute batch orders from Order Entry sheet
# ═══════════════════════════════════════════════════════════════════
def cmd_trade():
    print("T1 Broker — Executing Batch Orders...")
    wb, path = load_wb()
    client = get_client(wb)
    ws = wb["Order Entry"]

    # Collect orders with Execute=YES
    orders = []
    order_rows = []
    for r in range(5, 25):  # rows 5-24 (20 order slots)
        execute = str(ws.cell(row=r, column=9).value or "").strip().upper()
        symbol = str(ws.cell(row=r, column=2).value or "").strip()
        if execute != "YES" or not symbol:
            continue

        side = str(ws.cell(row=r, column=3).value or "").strip().upper()
        order_type = str(ws.cell(row=r, column=4).value or "MARKET").strip().upper()
        quantity = ws.cell(row=r, column=5).value
        limit_price = ws.cell(row=r, column=6).value
        stop_price = ws.cell(row=r, column=7).value
        tif = str(ws.cell(row=r, column=8).value or "DAY").strip().upper()

        if not side or side not in ("BUY", "SELL"):
            ws.cell(row=r, column=10).value = "FAILED"
            ws.cell(row=r, column=10).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="F85149")
            ws.cell(row=r, column=12).value = "Invalid side — must be BUY or SELL"
            ws.cell(row=r, column=13).value = now_str()
            continue

        if not quantity or float(quantity) <= 0:
            ws.cell(row=r, column=10).value = "FAILED"
            ws.cell(row=r, column=10).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="F85149")
            ws.cell(row=r, column=12).value = "Quantity must be > 0"
            ws.cell(row=r, column=13).value = now_str()
            continue

        order = {
            "symbol": symbol,
            "side": side.lower(),
            "orderType": order_type.lower(),
            "quantity": float(quantity),
            "timeInForce": tif.lower(),
        }
        if limit_price and order_type in ("LIMIT", "STOP_LIMIT"):
            order["price"] = float(limit_price)
        if stop_price and order_type in ("STOP", "STOP_LIMIT"):
            order["stopPrice"] = float(stop_price)

        orders.append(order)
        order_rows.append(r)

    if not orders:
        print("  No orders found with Execute=YES and valid symbol.")
        save_wb(wb, path)
        return

    print(f"  Submitting {len(orders)} orders as batch...")

    try:
        result = client.batch_orders(orders)
        results = result.get("results", [])

        for i, res in enumerate(results):
            r = order_rows[i]
            success = res.get("success", False)

            if success:
                ws.cell(row=r, column=10).value = "PLACED"
                ws.cell(row=r, column=10).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="3FB950")
                order_data = res.get("data", res)
                oid = order_data.get("id", res.get("id", "—"))
                ws.cell(row=r, column=11).value = oid
                ws.cell(row=r, column=11).font = openpyxl.styles.Font(name="Arial", size=10, color="58A6FF")
                status = order_data.get("status", "submitted")
                ws.cell(row=r, column=12).value = f"Order {status}"
            else:
                ws.cell(row=r, column=10).value = "FAILED"
                ws.cell(row=r, column=10).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="F85149")
                ws.cell(row=r, column=12).value = res.get("error", "Unknown error")

            ws.cell(row=r, column=12).font = openpyxl.styles.Font(name="Arial", size=10, color="8B949E")
            ws.cell(row=r, column=13).value = now_str()
            ws.cell(row=r, column=13).font = openpyxl.styles.Font(name="Arial", size=10, color="8B949E")

        placed = result.get("placed", sum(1 for r in results if r.get("success")))
        failed = result.get("failed", sum(1 for r in results if not r.get("success")))
        print(f"  Done: {placed} placed, {failed} failed out of {len(orders)}")

    except T1Error as e:
        print(f"  Batch submission error: {e}")
        for r in order_rows:
            ws.cell(row=r, column=10).value = "ERROR"
            ws.cell(row=r, column=10).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="F85149")
            ws.cell(row=r, column=12).value = str(e)
            ws.cell(row=r, column=13).value = now_str()

    save_wb(wb, path)

# ═══════════════════════════════════════════════════════════════════
# QUICK — Execute single order from Quick Trade sheet
# ═══════════════════════════════════════════════════════════════════
def cmd_quick():
    print("T1 Broker — Quick Trade...")
    wb, path = load_wb()
    client = get_client(wb)
    ws = wb["Quick Trade"]

    symbol = str(ws["C4"].value or "").strip()
    side = str(ws["C5"].value or "").strip().upper()
    order_type = str(ws["C6"].value or "MARKET").strip().upper()
    quantity = ws["C7"].value
    limit_price = ws["C8"].value
    stop_price = ws["C9"].value
    tif = str(ws["C10"].value or "DAY").strip().upper()

    if not symbol:
        print("  ERROR: No symbol specified in Quick Trade C4")
        return
    if not side or side not in ("BUY", "SELL"):
        print("  ERROR: Side must be BUY or SELL (C5)")
        return
    if not quantity or float(quantity) <= 0:
        print("  ERROR: Quantity must be > 0 (C7)")
        return

    print(f"  {side} {quantity} {symbol} ({order_type})...")

    try:
        if side == "BUY":
            result = client.buy(symbol, float(quantity), order_type=order_type.lower(),
                               price=float(limit_price) if limit_price else None,
                               time_in_force=tif.lower())
        else:
            result = client.sell(symbol, float(quantity), order_type=order_type.lower(),
                                price=float(limit_price) if limit_price else None,
                                time_in_force=tif.lower())

        order_data = result.get("data", result)
        ws["C13"].value = "PLACED"
        ws["C13"].font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="3FB950")
        ws["C14"].value = order_data.get("id", "—")
        ws["C14"].font = openpyxl.styles.Font(name="Arial", size=10, color="58A6FF")
        ws["C15"].value = order_data.get("price", "Market")
        ws["C16"].value = f"Order {order_data.get('status', 'submitted')}"
        ws["C16"].font = openpyxl.styles.Font(name="Arial", size=10, color="8B949E")
        ws["C17"].value = now_str()
        ws["C17"].font = openpyxl.styles.Font(name="Arial", size=10, color="8B949E")

        print(f"  Order placed: ID {ws['C14'].value}")

    except T1Error as e:
        ws["C13"].value = "FAILED"
        ws["C13"].font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="F85149")
        ws["C14"].value = "—"
        ws["C15"].value = "—"
        ws["C16"].value = str(e)
        ws["C16"].font = openpyxl.styles.Font(name="Arial", size=10, color="F85149")
        ws["C17"].value = now_str()
        print(f"  FAILED: {e}")

    save_wb(wb, path)

# ═══════════════════════════════════════════════════════════════════
# REFRESH — Update Positions + Order History sheets
# ═══════════════════════════════════════════════════════════════════
def cmd_refresh():
    print("T1 Broker — Refreshing Positions & Order History...")
    wb, path = load_wb()
    client = get_client(wb)

    # ── POSITIONS SHEET ──
    ws3 = wb["Positions"]

    try:
        positions = client.positions()
        pv = client.portfolio_value()

        # Summary row
        ws3.cell(row=4, column=3).value = pv["positions"]
        ws3.cell(row=4, column=3).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="39D2C0")
        ws3.cell(row=4, column=5).value = pv["marketValue"]
        ws3.cell(row=4, column=5).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="3FB950")
        ws3.cell(row=4, column=5).number_format = "$#,##0.00"
        ws3.cell(row=4, column=7).value = pv["costBasis"]
        ws3.cell(row=4, column=7).font = openpyxl.styles.Font(name="Arial", size=10, color="8B949E")
        ws3.cell(row=4, column=7).number_format = "$#,##0.00"

        pnl_color = "3FB950" if pv["unrealizedPnl"] >= 0 else "F85149"
        ws3.cell(row=4, column=9).value = pv["unrealizedPnl"]
        ws3.cell(row=4, column=9).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color=pnl_color)
        ws3.cell(row=4, column=9).number_format = "$#,##0.00;($#,##0.00);-"
        ws3.cell(row=4, column=11).value = pv["pnlPercent"] / 100 if pv["pnlPercent"] else 0
        ws3.cell(row=4, column=11).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color=pnl_color)
        ws3.cell(row=4, column=11).number_format = "0.0%"

        # Clear old data
        for r in range(7, 37):
            for c in range(2, 13):
                ws3.cell(row=r, column=c).value = None

        # Fill positions
        for i, p in enumerate(positions[:30]):
            r = 7 + i
            ws3.cell(row=r, column=2).value = p.get("symbol", "—")
            ws3.cell(row=r, column=2).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="58A6FF")
            ws3.cell(row=r, column=3).value = p.get("asset_class", p.get("assetClass", "—"))
            ws3.cell(row=r, column=4).value = (p.get("side", "long")).upper()

            qty = float(p.get("quantity", 0))
            ws3.cell(row=r, column=5).value = qty
            ws3.cell(row=r, column=5).number_format = "#,##0.####"

            avg = float(p.get("avg_cost", p.get("avgCost", 0)) or 0)
            ws3.cell(row=r, column=6).value = avg
            ws3.cell(row=r, column=6).number_format = "$#,##0.00"

            last = float(p.get("last_price", p.get("lastPrice", 0)) or 0)
            ws3.cell(row=r, column=7).value = last
            ws3.cell(row=r, column=7).number_format = "$#,##0.00"

            mv = float(p.get("marketValue", p.get("market_value", 0)) or qty * last)
            ws3.cell(row=r, column=8).value = mv
            ws3.cell(row=r, column=8).number_format = "$#,##0.00"

            upnl = float(p.get("unrealizedPnl", p.get("unrealized_pnl", 0)) or 0)
            ws3.cell(row=r, column=9).value = upnl
            ws3.cell(row=r, column=9).number_format = "$#,##0.00;($#,##0.00);-"
            color = "3FB950" if upnl >= 0 else "F85149"
            ws3.cell(row=r, column=9).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color=color)

            pnl_pct = float(p.get("pnlPercent", p.get("pnl_percent", 0)) or 0)
            ws3.cell(row=r, column=10).value = pnl_pct / 100 if abs(pnl_pct) > 1 else pnl_pct
            ws3.cell(row=r, column=10).number_format = "0.0%"
            ws3.cell(row=r, column=10).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color=color)

            ws3.cell(row=r, column=11).value = p.get("exchange", "—")
            ws3.cell(row=r, column=12).value = now_str()
            ws3.cell(row=r, column=12).font = openpyxl.styles.Font(name="Arial", size=10, color="8B949E")

        print(f"  Positions: {len(positions)} loaded, ${pv['marketValue']:,.2f} total, ${pv['unrealizedPnl']:+,.2f} P/L")

    except T1Error as e:
        print(f"  Positions error: {e}")

    # ── ORDER HISTORY SHEET ──
    ws4 = wb["Order History"]

    try:
        orders = client.list_orders(limit=50)

        for r in range(5, 55):
            for c in range(2, 12):
                ws4.cell(row=r, column=c).value = None

        for i, o in enumerate(orders[:50]):
            r = 5 + i
            ws4.cell(row=r, column=2).value = o.get("id", "—")
            ws4.cell(row=r, column=2).font = openpyxl.styles.Font(name="Arial", size=10, color="58A6FF")

            ws4.cell(row=r, column=3).value = o.get("symbol", o.get("instrument_symbol", "—"))
            ws4.cell(row=r, column=3).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="E6EDF3")

            side = o.get("side", "—").upper()
            ws4.cell(row=r, column=4).value = side
            ws4.cell(row=r, column=4).font = openpyxl.styles.Font(name="Arial", size=10, bold=True,
                                                                     color="3FB950" if side == "BUY" else "F85149")

            ws4.cell(row=r, column=5).value = o.get("order_type", o.get("orderType", "—")).upper()
            ws4.cell(row=r, column=6).value = float(o.get("quantity", 0))
            ws4.cell(row=r, column=6).number_format = "#,##0.####"

            price = o.get("price", o.get("filled_price", None))
            ws4.cell(row=r, column=7).value = float(price) if price else None
            ws4.cell(row=r, column=7).number_format = "$#,##0.00"

            status = o.get("status", "—").upper()
            status_colors = {"FILLED": "3FB950", "PENDING": "E3B341", "CANCELLED": "8B949E", "REJECTED": "F85149"}
            ws4.cell(row=r, column=8).value = status
            ws4.cell(row=r, column=8).font = openpyxl.styles.Font(name="Arial", size=10, bold=True,
                                                                     color=status_colors.get(status, "E6EDF3"))

            ws4.cell(row=r, column=9).value = o.get("time_in_force", o.get("timeInForce", "—")).upper()
            ws4.cell(row=r, column=10).value = o.get("created_at", o.get("createdAt", "—"))
            ws4.cell(row=r, column=10).font = openpyxl.styles.Font(name="Arial", size=10, color="8B949E")

        print(f"  Order History: {len(orders)} orders loaded")

    except T1Error as e:
        print(f"  Order history error: {e}")

    save_wb(wb, path)

# ═══════════════════════════════════════════════════════════════════
# INSTRUMENTS — Load available tickers into Instruments sheet
# ═══════════════════════════════════════════════════════════════════
def cmd_instruments():
    print("T1 Broker — Loading Instruments...")
    wb, path = load_wb()
    client = get_client(wb)
    ws5 = wb["Instruments"]

    search = str(ws5["C3"].value or "").strip()
    asset_class = str(ws5["E3"].value or "").strip()
    if asset_class and asset_class.upper() == "ALL":
        asset_class = None

    try:
        instruments = client.search_instruments(
            query=search if search else None,
            asset_class=asset_class.lower() if asset_class else None,
            limit=100,
        )

        for r in range(6, 106):
            for c in range(2, 10):
                ws5.cell(row=r, column=c).value = None

        for i, inst in enumerate(instruments[:100]):
            r = 6 + i
            ws5.cell(row=r, column=2).value = inst.get("id", "—")
            ws5.cell(row=r, column=2).font = openpyxl.styles.Font(name="Arial", size=10, color="8B949E")

            ws5.cell(row=r, column=3).value = inst.get("symbol", "—")
            ws5.cell(row=r, column=3).font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="58A6FF")

            ws5.cell(row=r, column=4).value = inst.get("name", "—")
            ws5.cell(row=r, column=4).font = openpyxl.styles.Font(name="Arial", size=10, color="E6EDF3")

            ac = inst.get("asset_class", inst.get("assetClass", "—"))
            ac_colors = {"equity": "3FB950", "crypto": "E3B341", "etf": "58A6FF", "forex": "BC8CFF", "option": "D29922"}
            ws5.cell(row=r, column=5).value = ac
            ws5.cell(row=r, column=5).font = openpyxl.styles.Font(name="Arial", size=10, bold=True,
                                                                     color=ac_colors.get(ac, "E6EDF3"))

            ws5.cell(row=r, column=6).value = inst.get("exchange", "—")

            last_price = inst.get("last_price", inst.get("lastPrice", None))
            ws5.cell(row=r, column=7).value = float(last_price) if last_price else None
            ws5.cell(row=r, column=7).number_format = "$#,##0.00"

            bid = inst.get("bid", inst.get("bid_price", None))
            ws5.cell(row=r, column=8).value = float(bid) if bid else None
            ws5.cell(row=r, column=8).number_format = "$#,##0.00"

            tradable = inst.get("is_tradable", inst.get("isTradable", True))
            ws5.cell(row=r, column=9).value = "YES" if tradable else "NO"
            ws5.cell(row=r, column=9).font = openpyxl.styles.Font(name="Arial", size=10, bold=True,
                                                                     color="3FB950" if tradable else "F85149")

        print(f"  Instruments: {len(instruments)} loaded" + (f" (filter: '{search}')" if search else ""))

    except T1Error as e:
        print(f"  Instruments error: {e}")

    save_wb(wb, path)

# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════
COMMANDS = {
    "connect": cmd_connect,
    "trade": cmd_trade,
    "quick": cmd_quick,
    "refresh": cmd_refresh,
    "instruments": cmd_instruments,
}

def cmd_all():
    cmd_connect()
    print()
    cmd_instruments()
    print()
    cmd_refresh()

COMMANDS["all"] = cmd_all

def main():
    if len(sys.argv) < 2 or sys.argv[1] in ("-h", "--help", "help"):
        print("T1 Broker Excel Trading Bridge v1.0")
        print("=" * 45)
        print()
        print("Usage: python t1_excel_bridge.py <command>")
        print()
        print("Commands:")
        print("  connect       Test API connection, update Connection sheet")
        print("  trade         Execute batch orders from Order Entry sheet")
        print("  quick         Execute single order from Quick Trade sheet")
        print("  refresh       Refresh positions + order history")
        print("  instruments   Load available instruments/tickers")
        print("  all           Full sync (connect + instruments + refresh)")
        print()
        print(f"Template: {TEMPLATE}")
        print("Requirements: pip install requests openpyxl")
        return

    cmd = sys.argv[1].lower()
    if cmd not in COMMANDS:
        print(f"Unknown command: '{cmd}'")
        print(f"Available: {', '.join(COMMANDS.keys())}")
        sys.exit(1)

    COMMANDS[cmd]()

if __name__ == "__main__":
    main()
